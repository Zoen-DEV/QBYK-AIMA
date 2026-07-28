"""Tests del parseo del sheet (bulk): columna `formato` multi-red y fuentes.

Cubre el mapeo formato → tipo_post/redes (historia y reel no existen en
LinkedIn), las fuentes por fila (`youtube_url` / `texto`) y la compatibilidad
con el encabezado viejo `formato_instagram`. El trigger de archivo por URL
(`archivo_url`) está diferido en el bulk; solo se testea la clasificación del
archivo remoto (plumbing conservado en `remote_file.py`).
"""
import csv
import io

import pytest

import networks
import sheets


def _csv_bytes(rows: list[dict]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8")


def _base_row(**overrides) -> dict:
    row = {
        "youtube_url": "https://www.youtube.com/watch?v=abc123",
        "texto": "",
        "tono": "",
        "objetivo": "",
        "tipo_medio": "imagen",
        "fuente_imagen": "higgsfield",
        "formato": "",
        "carrusel_slides": "",
        "idioma": "auto",
        "linkedin": "",
        "instagram": "",
        "facebook": "",
        "fecha_hora": "",
    }
    row.update(overrides)
    return row


# ── Matriz formato ↔ redes ────────────────────────────────────────────────────

def test_networks_for_format_matrix():
    all_nets = list(networks.NETWORKS)
    feed = ["linkedin", "instagram", "facebook"]  # los formatos de imagen no incluyen TikTok
    assert networks.networks_for_format("imagen-unica", all_nets) == feed
    assert networks.networks_for_format("carrusel", all_nets) == feed
    assert networks.networks_for_format("historia", all_nets) == ["instagram", "facebook"]
    # reel suma TikTok (video vertical) a Instagram y Facebook.
    assert networks.networks_for_format("reel", all_nets) == ["instagram", "facebook", "tiktok"]
    # Formato desconocido/vacío no agrega redes fuera de las de feed.
    assert networks.networks_for_format("", ["linkedin"]) == ["linkedin"]


# ── Columna formato en el sheet ───────────────────────────────────────────────

def test_formato_reel_omits_linkedin_with_warning():
    specs, warnings = sheets.parse_sheet(_csv_bytes([_base_row(formato="reel")]), "posts.csv")
    assert len(specs) == 1
    p = specs[0]["params"]
    assert p["formato"] == "reel"
    assert p["tipo_post"] == "reel"
    assert p["redes"] == ["instagram", "facebook"]
    assert any("no existe en linkedin" in w.lower() for w in warnings)


def test_formato_reel_only_linkedin_skips_row():
    row = _base_row(formato="reel", linkedin="sí", instagram="no", facebook="no")
    specs, warnings = sheets.parse_sheet(_csv_bytes([row]), "posts.csv")
    assert specs == []
    assert any("no aplica a ninguna" in w for w in warnings)


def test_formato_historia_video_maps_historia_formato():
    row = _base_row(formato="historia", tipo_medio="video")
    specs, _ = sheets.parse_sheet(_csv_bytes([row]), "posts.csv")
    p = specs[0]["params"]
    assert p["tipo_post"] == "historia"
    assert p["historia_formato"] == "video"
    assert p["redes"] == ["instagram", "facebook"]


def test_formato_historia_imagen_por_defecto():
    specs, _ = sheets.parse_sheet(_csv_bytes([_base_row(formato="historia")]), "posts.csv")
    assert specs[0]["params"]["historia_formato"] == "imagen"


def test_formato_carrusel_keeps_all_networks():
    specs, _ = sheets.parse_sheet(_csv_bytes([_base_row(formato="carrusel")]), "posts.csv")
    p = specs[0]["params"]
    assert p["formato"] == "carrusel"
    assert p["formato_instagram"] == "carrusel"
    assert p["tipo_post"] == "post"
    assert p["redes"] == ["linkedin", "instagram", "facebook"]


def test_legacy_formato_instagram_header_still_works():
    row = _base_row()
    del row["formato"]
    row["formato_instagram"] = "carrusel"
    specs, _ = sheets.parse_sheet(_csv_bytes([row]), "posts.csv")
    p = specs[0]["params"]
    assert p["formato"] == "carrusel"
    assert p["formato_instagram"] == "carrusel"


# ── Fuentes por fila (youtube_url / texto) ────────────────────────────────────

def test_texto_row_becomes_texto_source():
    row = _base_row(youtube_url="", texto="hola mundo")
    specs, warnings = sheets.parse_sheet(_csv_bytes([row]), "posts.csv")
    assert len(specs) == 1
    spec = specs[0]
    assert spec["source"] == "texto"
    assert spec["params"]["source_type"] == "texto"
    assert spec["upload_bytes"] == b"hola mundo"
    assert spec["label"] == "hola mundo"


def test_multiple_sources_prefers_youtube_with_warning():
    row = _base_row(texto="hola mundo")
    specs, warnings = sheets.parse_sheet(_csv_bytes([row]), "posts.csv")
    assert specs[0]["params"]["source_type"] == "youtube"
    assert any("se usa 'youtube_url'" in w for w in warnings)


def test_row_without_any_source_is_skipped():
    row = _base_row(youtube_url="")
    specs, warnings = sheets.parse_sheet(_csv_bytes([row]), "posts.csv")
    assert specs == []
    assert any("se omite" in w for w in warnings)


# ── Modelos de generación por fila ────────────────────────────────────────────

def test_modelo_columns_parse():
    row = _base_row(modelo_imagen="gpt_image_2", modelo_video="seedance_2_0",
                    modelo_voz="elevenlabs")
    specs, _ = sheets.parse_sheet(_csv_bytes([row]), "posts.csv")
    p = specs[0]["params"]
    assert p["modelo_imagen"] == "gpt_image_2"
    assert p["modelo_video"] == "seedance_2_0"
    assert p["modelo_voz"] == "elevenlabs"


def test_modelo_vacio_usa_default():
    specs, _ = sheets.parse_sheet(_csv_bytes([_base_row()]), "posts.csv")
    p = specs[0]["params"]
    assert p["modelo_imagen"] == ""
    assert p["modelo_video"] == ""
    assert p["modelo_voz"] == ""


def test_modelo_invalido_cae_al_default_con_warning():
    row = _base_row(modelo_imagen="dall-e-9")
    specs, warnings = sheets.parse_sheet(_csv_bytes([row]), "posts.csv")
    assert specs[0]["params"]["modelo_imagen"] == ""
    assert any("modelo_imagen" in w for w in warnings)


# ── Plantilla ─────────────────────────────────────────────────────────────────

def test_template_has_new_columns_and_parses_back():
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    data = sheets.build_template_xlsx()
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Posts"]
    headers = [c.value for c in ws[1]]
    assert headers == sheets.COLUMNS
    assert "formato" in headers
    assert "archivo_url" not in headers
    assert "formato_instagram" not in headers

    # Las filas de ejemplo de la propia plantilla deben parsear sin errores.
    specs, _ = sheets.parse_sheet(data, "aima-plantilla-posts.xlsx")
    assert len(specs) == len(sheets.EXAMPLE_ROWS)
    # La fila de ejemplo con formato=reel queda sin LinkedIn.
    reel = [s for s in specs if s["params"]["formato"] == "reel"]
    assert reel and reel[0]["params"]["redes"] == ["instagram", "facebook"]


# ── Clasificación del archivo remoto ─────────────────────────────────────────

def test_classify_source_by_extension_and_magic():
    import sys
    from pathlib import Path
    sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))
    import remote_file as rf

    assert rf.classify_source("nota.ogg", b"") == "audio"
    assert rf.classify_source("doc.pdf", b"") == "texto"
    assert rf.classify_source("", b"%PDF-1.7 ...") == "texto"
    assert rf.classify_source("", b"OggS....") == "audio"
    assert rf.classify_source("", b"ID3\x04...") == "audio"
    assert rf.classify_source("", "texto plano".encode()) == "texto"
    with pytest.raises(RuntimeError):
        rf.classify_source("cosa.bin", b"\x00\x01\x02\x03\x04\x05\x06\x07")
