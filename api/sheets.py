"""Plantilla .xlsx y parseo del sheet para la creación de posts en lote.

Una fila del sheet = un post. Las columnas mapean al dict `params` que consume
`run_pipeline` (ver `create_job` en app.py). Las cuentas de Blotato y el dry-run
NO van en el sheet: se eligen una sola vez en la UI de lote y el endpoint las
inyecta por fila. El máximo es MAX_ROWS filas (las demás se ignoran con aviso).
"""
from __future__ import annotations

import csv
import io
from datetime import date, datetime
from pathlib import Path

import model_catalog
import networks

# Orden de columnas = orden en la plantilla. Los encabezados deben coincidir con
# lo que lee parse_sheet (la lectura es case-insensitive).
COLUMNS = [
    "youtube_url",
    "texto",
    "tono",
    "objetivo",
    "tipo_medio",
    "fuente_imagen",
    "modelo_imagen",
    "template_set",
    "formato",
    "carrusel_slides",
    "duracion_video",
    "modelo_video",
    "modelo_voz",
    "idioma",
    "linkedin",
    "instagram",
    "facebook",
    "fecha_hora",
]

MAX_ROWS = 12

# Valores permitidos por columna enum (el vacío siempre vale = default/auto).
# `formato` aplica a TODAS las redes de la fila: una red que no soporta el formato
# se omite con warning (historia/reel no existen en LinkedIn — ver networks.py).
ALLOWED = {
    "tono": {"", "educativo", "inspiracional", "personal"},
    "objetivo": {"", "engagement", "awareness", "trafico"},
    "tipo_medio": {"imagen", "video"},
    "fuente_imagen": {"higgsfield", "template"},
    # Modelos de generación (vacío = default del .env). Catálogo curado en
    # model_catalog.py — misma fuente que los selectores del flujo individual.
    "modelo_imagen": {""} | set(model_catalog.IMAGE_MODELS),
    "modelo_video": {""} | set(model_catalog.VIDEO_MODELS),
    "modelo_voz": {""} | set(model_catalog.TTS_MODELS),
    "formato": {"imagen-unica", "carrusel", "historia", "reel"},
    "idioma": {"auto", "es", "en"},
    # linkedin/instagram/facebook son sí/no por red → se parsean aparte (_net_yes).
}

# Valores que se interpretan como "sí" / "no" en las columnas de red (case-insensitive).
NET_YES = {"sí", "si", "s", "yes", "y", "true", "1", "x", "✓"}
NET_NO = {"no", "n", "false", "0"}

# Redes que el sheet expone como columnas sí/no. TikTok NO está: es solo-individual
# (destino de reels desde la página /reel), así que el bulk nunca lo elige. Recorrer
# solo estas columnas evita que una columna ausente cuente como "sí".
SHEET_NETWORKS = ("linkedin", "instagram", "facebook")

# Opciones (ordenadas) que se muestran como lista desplegable en la plantilla.
# Deben mantenerse en sincronía con ALLOWED (sin el vacío: ese se deja en blanco).
# carrusel_slides también es una lista para que el usuario elija un número válido.
DROPDOWN_OPTIONS = {
    "tono": ["educativo", "inspiracional", "personal"],
    "objetivo": ["engagement", "awareness", "trafico"],
    "tipo_medio": ["imagen", "video"],
    "fuente_imagen": ["higgsfield", "template"],
    "modelo_imagen": list(model_catalog.IMAGE_MODELS),
    "modelo_video": list(model_catalog.VIDEO_MODELS),
    "modelo_voz": list(model_catalog.TTS_MODELS),
    "formato": ["imagen-unica", "carrusel", "historia", "reel"],
    "idioma": ["auto", "es", "en"],
    "carrusel_slides": ["3", "4", "5", "6"],
    "template_set": ["1", "2", "3"],
    "duracion_video": ["10", "20", "30", "45", "60"],
    "linkedin": ["sí", "no"],
    "instagram": ["sí", "no"],
    "facebook": ["sí", "no"],
}

DEFAULTS = {
    "tono": "",
    "objetivo": "",
    "tipo_medio": "imagen",
    "fuente_imagen": "higgsfield",
    "modelo_imagen": "",
    "template_set": 1,
    "modelo_video": "",
    "modelo_voz": "",
    "formato": "imagen-unica",
    "idioma": "auto",
    "carrusel_slides": 3,
    "duracion_video": 0,
    "linkedin": "sí",
    "instagram": "sí",
    "facebook": "sí",
}

# Descripción legible de valores válidos: va en los comentarios de celda y en la
# hoja "Instrucciones" de la plantilla.
COLUMN_HELP = {
    "youtube_url": "URL de YouTube. Llena ESTA o 'texto' (una sola fuente por fila).",
    "texto": "Texto libre (guion, notas). Llena ESTA o 'youtube_url' (una sola fuente por fila).",
    "tono": "Vacío (auto) | educativo | inspiracional | personal",
    "objetivo": "Vacío (auto) | engagement | awareness | trafico",
    "tipo_medio": "imagen | video. En formato = historia elige si la historia es imagen o video; en formato = reel se ignora (un reel siempre es video).",
    "fuente_imagen": "higgsfield (IA, con respaldo en plantillas) | template (solo plantillas). Solo aplica si tipo_medio = imagen.",
    "modelo_imagen": "Modelo de las imágenes IA. Vacío (default Nano Banana Pro) | nano_banana_pro (2 cr/img) | nano_banana_2 (1.5) | nano_banana (1) | gpt_image_2 (0.5) | z_image (0.15). Solo aplica si la fuente es higgsfield.",
    "template_set": "Set de estilo de las plantillas: 1 | 2 | 3 (vacío = 1). Solo aplica cuando se usan plantillas (fuente_imagen=template o como respaldo de Higgsfield).",
    "formato": "imagen-unica | carrusel | historia | reel. Aplica a todas las redes de la fila; si una red no soporta el formato se omite esa red (historia y reel no existen en LinkedIn). reel siempre genera video (requiere Higgsfield).",
    "carrusel_slides": "Número de 3 a 6 (solo aplica si formato = carrusel)",
    "duracion_video": "Duración del video en segundos: 10 | 20 | 30 | 45 | 60 (solo aplica si tipo_medio = video o formato = reel). Vacío = un solo clip corto. Se arma concatenando varios clips.",
    "modelo_video": "Modelo del video. Vacío (default Kling 3.0 Turbo) | kling3_0_turbo (1.5 cr/seg) | seedance_2_0_mini (2.5) | seedance_2_0 (4.5, máxima calidad). Solo aplica si hay video (tipo_medio = video o formato = reel/historia en video).",
    "modelo_voz": "Voz en off de los reels. Vacío (default Seed Audio) | seed_audio (~0.007 cr/carácter) | elevenlabs (~0.003). Solo aplica a videos con voz.",
    "idioma": "auto | es | en",
    "linkedin": "¿Publicar en LinkedIn? sí | no (vacío = sí)",
    "instagram": "¿Publicar en Instagram? sí | no (vacío = sí)",
    "facebook": "¿Publicar en Facebook? sí | no (vacío = sí)",
    "fecha_hora": "Programar el post, formato AAAA-MM-DD HH:MM. Vacío = publicar ahora.",
}

# Ancho de columna (en caracteres) para que la plantilla respire y se lea mejor.
COLUMN_WIDTHS = {
    "youtube_url": 44,
    "texto": 52,
    "tono": 18,
    "objetivo": 18,
    "tipo_medio": 16,
    "fuente_imagen": 18,
    "modelo_imagen": 20,
    "template_set": 14,
    "formato": 18,
    "carrusel_slides": 18,
    "duracion_video": 16,
    "modelo_video": 20,
    "modelo_voz": 16,
    "idioma": 14,
    "linkedin": 12,
    "instagram": 12,
    "facebook": 12,
    "fecha_hora": 22,
}

EXAMPLE_ROWS = [
    {
        "youtube_url": "https://www.youtube.com/watch?v=dQw4w9WgXcQ",
        "texto": "",
        "tono": "educativo",
        "objetivo": "engagement",
        "tipo_medio": "imagen",
        "fuente_imagen": "higgsfield",
        "modelo_imagen": "gpt_image_2",
        "formato": "carrusel",
        "carrusel_slides": 4,
        "duracion_video": "",
        "modelo_video": "",
        "modelo_voz": "",
        "idioma": "auto",
        "linkedin": "sí",
        "instagram": "sí",
        "facebook": "sí",
        "fecha_hora": "2026-06-20 09:00",
    },
    {
        "youtube_url": "",
        "texto": "Hoy quiero compartir 3 aprendizajes sobre productividad que cambiaron mi forma de trabajar...",
        "tono": "personal",
        "objetivo": "awareness",
        "tipo_medio": "imagen",
        "fuente_imagen": "template",
        "modelo_imagen": "",
        "formato": "imagen-unica",
        "carrusel_slides": 3,
        "duracion_video": "",
        "modelo_video": "",
        "modelo_voz": "",
        "idioma": "es",
        "linkedin": "sí",
        "instagram": "no",
        "facebook": "sí",
        "fecha_hora": "",
    },
    {
        "youtube_url": "https://www.youtube.com/watch?v=dQw4w9WgXcQ",
        "texto": "",
        "tono": "inspiracional",
        "objetivo": "awareness",
        "tipo_medio": "video",
        "fuente_imagen": "higgsfield",
        "modelo_imagen": "",
        "formato": "reel",
        "carrusel_slides": 3,
        "duracion_video": "30",
        "modelo_video": "seedance_2_0",
        "modelo_voz": "elevenlabs",
        "idioma": "auto",
        "linkedin": "no",
        "instagram": "sí",
        "facebook": "sí",
        "fecha_hora": "",
    },
]

# Formatos de fecha aceptados cuando la celda viene como texto (no como fecha de Excel).
_DATE_FORMATS = [
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%dT%H:%M",
    "%d/%m/%Y %H:%M:%S",
    "%d/%m/%Y %H:%M",
    "%Y-%m-%d",
    "%d/%m/%Y",
]


# ── Plantilla ────────────────────────────────────────────────────────────────────

def build_template_xlsx() -> bytes:
    """Genera la plantilla .xlsx descargable (encabezados + ejemplos + instrucciones)."""
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "Posts"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="7C3AED")  # violeta marca
    example_font = Font(italic=True, color="6B7280")       # gris: filas de ejemplo
    example_fill = PatternFill("solid", fgColor="F3EFFF")  # violeta muy claro
    thin = Side(style="thin", color="D1C7EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Las columnas de texto libre se alinean a la izquierda; el resto al centro.
    left_cols = {"youtube_url", "texto"}
    # La lista desplegable cubre exactamente las filas que el lote procesa.
    last_row = 1 + MAX_ROWS

    # Encabezados.
    for ci, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        cell.comment = Comment(COLUMN_HELP[col], "AIMA")
        ws.column_dimensions[get_column_letter(ci)].width = COLUMN_WIDTHS.get(col, 18)
    ws.row_dimensions[1].height = 30

    # Filas de ejemplo (estilo tenue para que se note que se pueden borrar).
    for ri, ex in enumerate(EXAMPLE_ROWS, start=2):
        for ci, col in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=ri, column=ci, value=ex.get(col, ""))
            cell.font = example_font
            cell.fill = example_fill
            cell.border = border
            align = "left" if col in left_cols else "center"
            cell.alignment = Alignment(horizontal=align, vertical="center")
        ws.row_dimensions[ri].height = 22

    # Filas vacías restantes: solo borde y alineación, listas para escribir.
    for ri in range(2 + len(EXAMPLE_ROWS), last_row + 1):
        for ci, col in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=ri, column=ci)
            cell.border = border
            align = "left" if col in left_cols else "center"
            cell.alignment = Alignment(horizontal=align, vertical="center")
        ws.row_dimensions[ri].height = 22

    # Listas desplegables (selects) en las columnas de opciones fijas: evitan typos.
    for col, options in DROPDOWN_OPTIONS.items():
        letter = get_column_letter(COLUMNS.index(col) + 1)
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(options) + '"',
            allow_blank=True,          # el vacío = default/auto sigue siendo válido
            showErrorMessage=True,
        )
        dv.errorTitle = "Valor no válido"
        dv.error = f"Elige una opción de la lista: {', '.join(options)}."
        dv.add(f"{letter}2:{letter}{last_row}")
        ws.add_data_validation(dv)

    ws.freeze_panes = "A2"

    # Hoja de instrucciones
    ws2 = wb.create_sheet("Instrucciones")
    intro = [
        ("Cómo usar esta plantilla", True),
        ("", False),
        (f"• Cada fila = un post. Máximo {MAX_ROWS} filas (las demás se ignoran).", False),
        ("• Por fila llena UNA fuente: 'youtube_url' o 'texto' (texto libre como guion o notas).", False),
        ("• 'fecha_hora' programa el post (AAAA-MM-DD HH:MM). Vacío = publicar ahora.", False),
        ("• Redes: pon sí/no en las columnas linkedin, instagram y facebook. Vacío = sí (se publica en las tres por defecto).", False),
        ("• 'formato' aplica a todas las redes de la fila. historia y reel no existen en LinkedIn: esa red se omite y se publica en las demás.", False),
        ("• Las columnas con lista desplegable solo aceptan los valores de la lista.", False),
        ("• No borres ni renombres la fila de encabezados (fila 1).", False),
        ("• Las cuentas de LinkedIn/Instagram/Facebook se eligen en la app, no en el sheet.", False),
        ("• Las filas de ejemplo (en gris) son de muestra: puedes borrarlas antes de cargar.", False),
        ("", False),
        ("Valores válidos por columna", True),
        ("", False),
    ]
    for ri, (text, bold) in enumerate(intro, start=1):
        c = ws2.cell(row=ri, column=1, value=text)
        if bold:
            c.font = Font(bold=True, size=13)
    base = len(intro) + 1
    for i, col in enumerate(COLUMNS):
        ws2.cell(row=base + i, column=1, value=f"{col}: {COLUMN_HELP[col]}")
    ws2.column_dimensions["A"].width = 95

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Parseo ───────────────────────────────────────────────────────────────────────

def parse_sheet(data: bytes, filename: str = "") -> tuple[list[dict], list[str]]:
    """Lee un .xlsx/.csv y devuelve (specs, warnings).

    Cada spec es un dict con: params (normalizado, sin cuentas/dry-run), source,
    label, schedule_dt (naive o None), upload_bytes, upload_filename.
    Lanza ValueError si el archivo no se puede leer.
    """
    warnings: list[str] = []
    fmt = _detect_format(data, filename)
    try:
        raw_rows = _rows_from_csv(data) if fmt == "csv" else _rows_from_xlsx(data)
    except Exception as e:  # noqa: BLE001 - se reporta al usuario
        raise ValueError(f"No se pudo leer el archivo ({fmt}): {e}")

    # Descarta filas totalmente vacías.
    raw_rows = [r for r in raw_rows if any(_clean(v) for v in r.values())]

    if len(raw_rows) > MAX_ROWS:
        warnings.append(
            f"El archivo tiene {len(raw_rows)} filas; se procesan solo las primeras {MAX_ROWS}."
        )
        raw_rows = raw_rows[:MAX_ROWS]

    specs: list[dict] = []
    for idx, r in enumerate(raw_rows, start=1):
        spec, row_warnings = _row_to_spec(r, idx)
        warnings.extend(row_warnings)
        if spec is not None:
            specs.append(spec)

    return specs, warnings


def _detect_format(data: bytes, filename: str) -> str:
    ext = Path(filename or "").suffix.lower()
    if ext == ".csv":
        return "csv"
    if ext in (".xlsx", ".xlsm"):
        return "xlsx"
    # Sin extensión confiable: los .xlsx son contenedores ZIP (PK\x03\x04).
    return "xlsx" if data[:4] == b"PK\x03\x04" else "csv"


def _rows_from_csv(data: bytes) -> list[dict]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    rows: list[dict] = []
    for row in reader:
        rows.append({(k or "").strip(): v for k, v in row.items()})
    return rows


def _rows_from_xlsx(data: bytes) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    headers = [str(h).strip() if h is not None else "" for h in header]
    out: list[dict] = []
    for row in rows_iter:
        d: dict = {}
        for i, h in enumerate(headers):
            if h:
                d[h] = row[i] if i < len(row) else None
        out.append(d)
    return out


def _row_to_spec(r: dict, idx: int) -> tuple[dict | None, list[str]]:
    w: list[str] = []
    g = {(k or "").strip().lower(): v for k, v in r.items()}

    youtube_url = _clean(g.get("youtube_url"))
    texto = _clean(g.get("texto"))

    filled = [c for c, v in (("youtube_url", youtube_url), ("texto", texto)) if v]
    if len(filled) > 1:
        w.append(f"Fila {idx}: tiene {', '.join(filled)}; se usa '{filled[0]}'.")
        if filled[0] != "texto":
            texto = ""
    if not filled:
        w.append(f"Fila {idx}: sin 'youtube_url' ni 'texto'; se omite.")
        return None, w

    source = "youtube" if youtube_url else "texto"
    upload_bytes = b""
    upload_filename = ""
    if source == "texto":
        upload_bytes = texto.encode("utf-8")
        upload_filename = "texto.txt"

    # `formato` aplica a todas las redes (se acepta el encabezado viejo
    # 'formato_instagram' por compatibilidad con plantillas ya descargadas).
    formato_cell = g.get("formato")
    if _clean(formato_cell) == "" and _clean(g.get("formato_instagram")):
        formato_cell = g.get("formato_instagram")
    formato = _enum(formato_cell, "formato", w, idx)
    tipo_medio = _enum(g.get("tipo_medio"), "tipo_medio", w, idx)

    # historia/reel se modelan como tipo_post (el mismo discriminador del flujo
    # individual); el formato de feed que consume el pipeline queda en
    # formato_instagram. reel siempre es video; historia usa tipo_medio.
    tipo_post = formato if formato in ("historia", "reel") else "post"
    historia_formato = "video" if (formato == "historia" and tipo_medio == "video") else "imagen"

    redes = _parse_net_flags(g, w, idx)
    redes_ok = networks.networks_for_format(formato, redes)
    if not redes_ok:
        w.append(
            f"Fila {idx}: el formato '{formato}' no aplica a ninguna de las redes elegidas; se omite la fila."
        )
        return None, w
    dropped = [n for n in redes if n not in redes_ok]
    if dropped:
        w.append(
            f"Fila {idx}: el formato '{formato}' no existe en {', '.join(dropped)}; esa red se omite."
        )

    params = {
        "source_type": source,
        "youtube_url": youtube_url,
        "upload_filename": upload_filename,
        "tono": _enum(g.get("tono"), "tono", w, idx),
        "tono_linkedin": "",
        "tono_instagram": "",
        "tono_facebook": "",
        "objetivo": _enum(g.get("objetivo"), "objetivo", w, idx),
        "objetivo_linkedin": "",
        "objetivo_instagram": "",
        "objetivo_facebook": "",
        "formato": formato,
        "formato_instagram": "carrusel" if formato == "carrusel" else "imagen-unica",
        "tipo_post": tipo_post,
        "media_origin": "generar",
        "historia_formato": historia_formato,
        "carrusel_slides": _parse_slides(g.get("carrusel_slides"), w, idx),
        "duracion_video": _parse_duracion(g.get("duracion_video"), w, idx),
        "camara_estilo": "dolly",
        "tipo_medio": tipo_medio,
        "fuente_imagen": _enum(g.get("fuente_imagen"), "fuente_imagen", w, idx),
        "modelo_imagen": _enum(g.get("modelo_imagen"), "modelo_imagen", w, idx),
        "template_set": _parse_template_set(g.get("template_set"), w, idx),
        "modelo_video": _enum(g.get("modelo_video"), "modelo_video", w, idx),
        "modelo_voz": _enum(g.get("modelo_voz"), "modelo_voz", w, idx),
        "idioma": _enum(g.get("idioma"), "idioma", w, idx),
        "modelo_perplexity": "sonar-pro",
        "redes": redes_ok,
        "publicar": "",
    }

    schedule_dt = _parse_datetime(g.get("fecha_hora"), w, idx)
    if source == "youtube":
        label = youtube_url
    else:
        label = texto[:80] + ("…" if len(texto) > 80 else "")

    spec = {
        "params": params,
        "source": source,
        "label": label,
        "schedule_dt": schedule_dt,
        "upload_bytes": upload_bytes,
        "upload_filename": upload_filename,
    }
    return spec, w


# ── Helpers de normalización ──────────────────────────────────────────────────────

def _clean(v) -> str:
    """Valor de celda → string limpio (sin convertir fechas, que se manejan aparte)."""
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, (datetime, date)):
        return v.isoformat(sep=" ")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _enum(value, col: str, w: list[str], idx: int) -> str:
    v = _clean(value).lower()
    if v in ALLOWED[col]:
        return v
    # Celda vacía = usar el default en silencio (no es un error del usuario).
    if v == "":
        return DEFAULTS[col]
    w.append(f"Fila {idx}: '{col}' inválido ('{_clean(value)}'); se usa el valor por defecto.")
    return DEFAULTS[col]


def _net_yes(value, net: str, w: list[str], idx: int) -> bool:
    """Celda sí/no de una columna de red → bool. Vacío = sí (publicar por defecto)."""
    v = _clean(value).lower()
    if v == "" or v in NET_YES:
        return True
    if v in NET_NO:
        return False
    w.append(f"Fila {idx}: valor de '{net}' no reconocido ('{_clean(value)}'); se incluye la red.")
    return True


def _parse_net_flags(g: dict, w: list[str], idx: int) -> list[str]:
    """Columnas linkedin/instagram/facebook (sí/no) → lista canónica de redes.

    Por defecto (todas vacías) se publican las tres. Si el usuario pone 'no' en las
    tres, se avisa y se vuelve al default para no dejar el post sin destino.
    """
    chosen = [net for net in SHEET_NETWORKS if _net_yes(g.get(net), net, w, idx)]
    if not chosen:
        w.append(f"Fila {idx}: no se eligió ninguna red (todas en 'no'); se publican las tres.")
        return list(SHEET_NETWORKS)
    return chosen


def _parse_slides(value, w: list[str], idx: int) -> int:
    s = _clean(value)
    if not s:
        return DEFAULTS["carrusel_slides"]
    try:
        n = int(float(s))
    except (ValueError, TypeError):
        w.append(f"Fila {idx}: 'carrusel_slides' inválido ('{s}'); se usa 3.")
        return DEFAULTS["carrusel_slides"]
    clamped = max(3, min(6, n))
    if clamped != n:
        w.append(f"Fila {idx}: 'carrusel_slides' {n} fuera de rango; se ajusta a {clamped}.")
    return clamped


def _parse_template_set(value, w: list[str], idx: int) -> int:
    """Set de estilo de plantillas (1-3). Vacío → 1; fuera de rango → 1 con aviso."""
    s = _clean(value)
    if not s:
        return DEFAULTS["template_set"]
    try:
        n = int(float(s))
    except (ValueError, TypeError):
        w.append(f"Fila {idx}: 'template_set' inválido ('{s}'); se usa 1.")
        return DEFAULTS["template_set"]
    if n not in (1, 2, 3):
        w.append(f"Fila {idx}: 'template_set' {n} fuera de rango (1-3); se usa 1.")
        return DEFAULTS["template_set"]
    return n


def _parse_duracion(value, w: list[str], idx: int) -> int:
    """Duración de video en segundos (0 = un solo clip corto por defecto).

    Vacío → 0. Se clampa a 0–60 (el pipeline la logra concatenando varios clips).
    """
    s = _clean(value)
    if not s:
        return DEFAULTS["duracion_video"]
    try:
        n = int(float(s))
    except (ValueError, TypeError):
        w.append(f"Fila {idx}: 'duracion_video' inválido ('{s}'); se usa el clip por defecto.")
        return DEFAULTS["duracion_video"]
    clamped = max(0, min(60, n))
    if clamped != n:
        w.append(f"Fila {idx}: 'duracion_video' {n} fuera de rango; se ajusta a {clamped}.")
    return clamped


def _parse_datetime(value, w: list[str], idx: int) -> datetime | None:
    """Celda fecha/hora → datetime naive (hora local del usuario) o None si vacía."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    s = str(value).strip()
    if not s:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    w.append(f"Fila {idx}: 'fecha_hora' no reconocida ('{s}'); ese post se publicará de inmediato.")
    return None
